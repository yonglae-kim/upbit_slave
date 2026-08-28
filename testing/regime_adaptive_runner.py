#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.8"
# dependencies = ["openpyxl"]
# ///

# ─── How to run ───
# 1. Install uv (if not installed): curl -LsSf https://astral.sh/uv/install.sh | sh
# 2. Run: uv run testing/regime_adaptive_runner.py --help
# ──────────────────

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from zipfile import BadZipFile
from datetime import datetime, timedelta, timezone
from math import isfinite
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple, Union

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from testing.cross_sectional_contrarian_evaluation import validate_manifest
from testing.cross_sectional_momentum_evaluation import (
    MARKETS,
    AlignedMarketData,
    Candle,
    EvaluationInputError,
    align_series,
)
from testing.regime_adaptive_evaluation import run_evaluation
from testing.regime_adaptive_models import (
    EvaluationWindow,
    RegimeAdaptiveConfig,
    RegimeAdaptiveResult,
)

JsonValue = Union[str, int, float, bool, None, List["JsonValue"], Dict[str, "JsonValue"]]


def _parse_utc(raw: str) -> datetime:
    """Parse a timezone-explicit UTC ISO-8601 timestamp."""
    value = raw.strip()
    if value.endswith("Z"):
        value = value[:-1] + "+00:00"
    elif not value.endswith("+00:00"):
        raise EvaluationInputError("timestamp must be explicit UTC ISO-8601")
    try:
        parsed = datetime.fromisoformat(value)
    except (TypeError, ValueError) as error:
        raise EvaluationInputError("timestamp must be explicit UTC ISO-8601") from error
    if parsed.tzinfo is None or parsed.utcoffset() != timedelta(0):
        raise EvaluationInputError("timestamp must be explicit UTC ISO-8601")
    return parsed.astimezone(timezone.utc)


def _load_market(market: str, source: Path) -> Tuple[str, Tuple[Candle, ...]]:
    """Load and validate real OHLC rows from one local XLSX workbook."""
    try:
        workbook = load_workbook(filename=str(source), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as error:
        raise EvaluationInputError("invalid workbook: " + market) from error
    candles: List[Candle] = []
    try:
        iterator = workbook.active.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration as error:
            raise EvaluationInputError("workbook has no header: " + market) from error
        headers = tuple("" if value is None else str(value).strip() for value in raw_headers)
        required = ("candle_date_time_utc", "opening_price", "high_price", "low_price", "trade_price")
        if any(headers.count(name) != 1 for name in required):
            raise EvaluationInputError("workbook columns are invalid: " + market)
        positions = {name: headers.index(name) for name in required}
        row_error: Optional[str] = None
        for row in iterator:
            try:
                raw_timestamp = row[positions["candle_date_time_utc"]]
                if not isinstance(raw_timestamp, str):
                    raise EvaluationInputError("candle timestamp is not explicit UTC: " + market)
                timestamp = _parse_utc(raw_timestamp)
                opening = float(row[positions["opening_price"]])
                high = float(row[positions["high_price"]])
                low = float(row[positions["low_price"]])
                close = float(row[positions["trade_price"]])
            except (IndexError, TypeError, ValueError, OverflowError) as error:
                if row_error is None:
                    row_error = "invalid candle row: " + market
                continue
            values = (opening, high, low, close)
            if any(not isfinite(value) or value <= 0.0 for value in values):
                if row_error is None:
                    row_error = "candle values must be finite and positive: " + market
                continue
            if low > high or low > opening or low > close or high < opening or high < close:
                if row_error is None:
                    row_error = "candle prices are inverted: " + market
                continue
            candles.append(Candle(timestamp, close, low))
    finally:
        iterator.close()
        workbook.close()
        del iterator
        del workbook
    if row_error is not None:
        raise EvaluationInputError(row_error)
    ordered = tuple(sorted(candles, key=lambda candle: candle.timestamp))
    if not ordered:
        raise EvaluationInputError("workbook has no candle rows: " + market)
    if any(left.timestamp == right.timestamp for left, right in zip(ordered, ordered[1:])):
        raise EvaluationInputError("duplicate candle timestamp: " + market)
    return market, ordered


def _load_data(manifest: Path) -> AlignedMarketData:
    """Validate the manifest, load all eight local XLSX files, and align candles."""
    entries = validate_manifest(manifest)
    if len(entries) != len(MARKETS) or any(source.suffix.lower() != ".xlsx" for _, source in entries):
        raise EvaluationInputError("manifest must declare eight local XLSX files")
    return align_series(tuple(_load_market(market, source) for market, source in entries))


def _iso(value: datetime) -> str:
    """Serialize a UTC timestamp with the canonical Z suffix."""
    return value.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def _sha256(path: Path) -> str:
    """Return the SHA-256 digest of the manifest bytes."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _parser() -> argparse.ArgumentParser:
    """Build the regime-adaptive evaluation CLI parser."""
    parser = argparse.ArgumentParser(description="Evaluation-only regime-adaptive backtest")
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--window-start", required=True)
    parser.add_argument("--window-end", required=True)
    parser.add_argument("--regime-lookback", type=int, default=288)
    parser.add_argument("--bull-threshold", type=float, default=0.03)
    parser.add_argument("--bear-threshold", type=float, default=-0.03)
    parser.add_argument("--momentum-lookback", type=int, default=36)
    parser.add_argument("--hold-bars", type=int, default=48)
    parser.add_argument("--breadth-min", type=int, default=5)
    parser.add_argument("--selected-return-max", type=float, default=-0.005)
    parser.add_argument("--stop-loss-pct", type=float, default=0.03)
    parser.add_argument("--exposure", type=float, default=0.50)
    parser.add_argument("--fee", type=float, default=0.0005)
    parser.add_argument("--spread", type=float, default=0.0003)
    parser.add_argument("--slippage", type=float, default=0.0002)
    return parser


def _payload(
    manifest: Path, data: AlignedMarketData, config: RegimeAdaptiveConfig,
    window: EvaluationWindow, result: RegimeAdaptiveResult,
) -> Dict[str, JsonValue]:
    """Convert the evaluator result to the stable JSON contract."""
    return {
        "evaluation_only": True,
        "window_convention": "half_open",
        "window_start": _iso(window.start),
        "window_end": _iso(window.end),
        "markets": list(data.markets),
        "strategy_by_regime": {
            "bull": "cross_sectional_momentum",
            "sideways": "cross_sectional_contrarian_bounce",
            "bear": "cash_defensive",
        },
        "parameters": {
            "regime_lookback": config.regime_lookback,
            "bull_threshold": config.bull_threshold,
            "bear_threshold": config.bear_threshold,
            "momentum_lookback": config.momentum_lookback,
            "hold_bars": config.hold_bars,
            "breadth_min": config.breadth_min,
            "selected_return_max": config.selected_return_max,
            "stop_loss_pct": config.stop_loss_pct,
            "exposure": config.exposure,
        },
        "cost_model": {
            "fee": config.fee,
            "spread": config.spread,
            "slippage": config.slippage,
            "one_side_cost": config.one_side_cost,
        },
        "regime_counts": {
            item.regime.value: {"labels": item.labels, "trades": item.trades, "portfolio_return": item.portfolio_return}
            for item in result.regime_counts
        },
        "strategy_summaries": {
            item.strategy.value: {
                "trades": item.trades,
                "winning_trades": item.winning_trades,
                "portfolio_return": item.portfolio_return,
            }
            for item in result.strategy_summaries
        },
        "total_return": result.total_return,
        "max_drawdown": result.max_drawdown,
        "trades_detail": [
            {
                "market": trade.market,
                "regime": trade.regime.value,
                "strategy": trade.strategy.value,
                "entry_time": _iso(trade.entry_time),
                "exit_time": _iso(trade.exit_time),
                "entry_price": trade.entry_price,
                "exit_price": trade.exit_price,
                "gross_return": trade.gross_return,
                "net_return": trade.net_return,
                "portfolio_return": trade.portfolio_return,
                "btc_return": trade.btc_return,
                "breadth": trade.breadth,
                "exit_reason": trade.exit_reason,
            }
            for trade in result.trade_records
        ],
        "manifest_sha256": _sha256(manifest),
        "integrity": {
            "fixed_universe": data.markets == MARKETS,
            "common_timestamp_rows": len(data.timestamps),
            "synthetic_rows": 0,
            "forward_fill": False,
            "real_ohlc_low_stop": True,
        },
    }


def main(argv: Optional[Sequence[str]] = None) -> int:
    """Run the offline evaluation and emit identical JSON to stdout and disk."""
    args = _parser().parse_args(argv)
    start = _parse_utc(args.window_start)
    end = _parse_utc(args.window_end)
    window = EvaluationWindow(start, end)
    config = RegimeAdaptiveConfig(
        regime_lookback=args.regime_lookback,
        bull_threshold=args.bull_threshold,
        bear_threshold=args.bear_threshold,
        momentum_lookback=args.momentum_lookback,
        hold_bars=args.hold_bars,
        breadth_min=args.breadth_min,
        selected_return_max=args.selected_return_max,
        stop_loss_pct=args.stop_loss_pct,
        exposure=args.exposure,
        fee=args.fee,
        spread=args.spread,
        slippage=args.slippage,
    )
    if config.one_side_cost >= 1.0:
        raise EvaluationInputError("combined cost must be less than one")
    manifest = args.manifest.resolve()
    data = _load_data(manifest)
    result = run_evaluation(data, config, window)
    if not isfinite(result.total_return) or not isfinite(result.max_drawdown):
        raise EvaluationInputError("evaluation metrics must be finite")
    serialized = json.dumps(_payload(manifest, data, config, window, result), indent=2, sort_keys=True, allow_nan=False)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(serialized, encoding="utf-8")
    print(serialized)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (EvaluationInputError, OSError) as error:
        print("error: " + str(error), file=sys.stderr)
        raise SystemExit(1)
